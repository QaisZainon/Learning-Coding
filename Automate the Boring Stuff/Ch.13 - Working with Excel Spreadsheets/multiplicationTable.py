#! python3
# Takes a number N from command line and creates a 
# N * N multiplication table in an Excel spreadsheet
import openpyxl
from openpyxl.styles import Font
# Defining the numbers for the chart
num_chart = [1, 2, 3, 4, 5, 6]
def multiplication(num):
    # Opens Workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Multiplication table'
    Bold = Font(bold = True)
    # Defining the variable of the table
    for i in range(1, len(num) +1):
        sheet.cell(row = 1, column = i + 1).value = i
        sheet.cell(row = 1, column = i + 1).font = Bold
        sheet.cell(row = i + 1, column = 1).value = i
        sheet.cell(row = i + 1, column = 1).font = Bold
    for row in range(2, len(num) +2):
        for column in range(2, len(num) +2):
            answer = num[(row - 2)] * num[(column - 2)]
            sheet.cell(row = row, column = column).value = answer
    wb.save('multiplicationTable.xlsx')

multiplication(num_chart)
