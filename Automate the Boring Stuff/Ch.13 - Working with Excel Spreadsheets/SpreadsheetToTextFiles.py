#! python3
# Open spreadsheets and writes the cells into one text file for each columns
import openpyxl 
wb = openpyxl.load_workbook('TextToExcel.xlsx')
sheet = wb.active
column_len = sheet.max_column
row_len = sheet.max_row
# Harvesting data
data_bank = {}
for column in range(1, column_len + 1):
    data_bank.update({column:[]})
    for row in range(1, row_len + 1):
        data_bank[column].append(sheet.cell(row = row, column = column).value)
# Placing the data into .txt files
for key in data_bank.keys():
    open_txt = open('num' + str(key) + '.txt', 'a')
    txt = ''
    txt = ' '.join(data_bank.get(key))
    open_txt.write(txt)
    open_txt.close()
