#! python3
# Inverts the cells and rows of a given table
import openpyxl
wb = openpyxl.load_workbook('SpreadsheetCellInverter.xlsx')
sheet = wb.active
column_len = sheet.max_column
row_len = sheet.max_row
data_bank = {}
# Saving the values into a list
for column in range(column_len):
    data_bank.update({column:[]})
    for row in range(row_len):
        data_bank[column].append(sheet[row + 1][column].value)
# Inverting it into new a table
wb = openpyxl.Workbook()
sheet = wb.active
for column in range(row_len):
    for row in range(column_len):
        sheet.cell(row = row + 1, column = column + 1).value = data_bank[row][column]
# Legit I dont know how i did this lmao
wb.save('SpreadsheetCellInverter_copy.xlsx')