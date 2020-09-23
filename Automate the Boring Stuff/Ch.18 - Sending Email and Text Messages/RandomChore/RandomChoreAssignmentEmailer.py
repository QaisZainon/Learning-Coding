#! python3
# RandomChoreAssignmentEmailer.py - Assigns chores to random people
import openpyxl, smtplib, random
email = 'qaisbinzainonwork@gmail.com'
email_pw = input('Input password for %s\n' % (email))
wb = openpyxl.load_workbook('Automate the Boring Stuff\Ch.18 - Sending Email and Text Messages\RandomChore\RandomChoices.xlsx')
sheet = wb['Sheet1']
maxColumn = sheet.max_column
maxRow = sheet.max_row
chores = ['Dishes', 'Bathroom', 'Vacuum', 'Walkdog']

# Extablishing SMTP connection
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(email, email_pw)

for r in range(2, maxRow + 1):
    randomChore = random.choice(chores)
    chores.remove(randomChore)  # this chore is now taken, so remove it
    recipient_name = sheet.cell(row = r, column = 1).value
    recipient_mail = sheet.cell(row = r, column = 2).value
    sheet.cell(row = r, column = 3).value = randomChore
    print('Sending %s with chore: %s,at mail: %s' % (recipient_name, randomChore, recipient_mail))
    #smtpObj.sendmail(email, recipient_mail, 'Subject: Chore Assignment for %s\nYou are assigned to doing %s' % (recipient_name, randomChore))

wb.save('Automate the Boring Stuff\Ch.18 - Sending Email and Text Messages\RandomChore\RandomChoices.xlsx')
smtpObj.quit()